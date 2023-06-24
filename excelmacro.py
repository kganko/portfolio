import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

wb = openpyxl.load_workbook("winequality-red1.xlsx")
sheet = wb["winequality-red"]
sheet.insert_cols(idx=1)
sheet['A1'] = 'my assessment'
sheet['O1'] = 'comments'
sheet.freeze_panes = 'A2'


sheet['A1'].font = Font(color="FF0000", italic=True)


sheet.page_setup.fitToWidth = 1
sheet.row_dimensions[1].height = 80


for rowNum in range(2, sheet.max_row+1):
    myAssessment = sheet.cell(row=rowNum, column=1).value
    sheet.cell(row=rowNum, column=1).value = 2


# conditional formatting
redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
greenFill = PatternFill(start_color='4EBC97', end_color='4EBC97', fill_type='solid')
sheet.conditional_formatting.add('B2:B1500', ColorScaleRule(start_type='min', start_color='AA0000', end_type='max', end_color='DD9797'))
sheet.conditional_formatting.add('C2:C1500', ColorScaleRule(start_type='percentile', start_value=10, start_color='4B75C5', mid_type='percentile', mid_value=50, mid_color='7495D2', end_type='percentile', end_value=90, end_color='BBCAE8'))
sheet.conditional_formatting.add('D2:D1500',  CellIsRule(operator='lessThan', formula=['0.4'], stopIfTrue=True, fill=redFill))
sheet.conditional_formatting.add('E2:E1500',CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=greenFill))
sheet.conditional_formatting.add('G1:G1500', FormulaRule(formula=['ISBLANK(G1)'], stopIfTrue=True, fill=redFill))

import pandas as pd
df = pd.read_excel("winequality-red1.xlsx")
df.pivot_table(index = ["fixed acidity", "volatile acidity", "citric acid", "residual sugar", "chlorides", "total sulfur dioxide"],
              aggfunc = [sum],
              margins=True,
              margins_name='Grand Totals')


wb.save("winequality-red.xlsx")

